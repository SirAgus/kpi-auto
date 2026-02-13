import os, io, json, requests, pandas as pd
from datetime import timedelta
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import msal
import re
import time

try:
    from groq import Groq
except ImportError:
    Groq = None

slack_bot_token=os.environ["SLACK_BOT_TOKEN"]
channel_id=os.environ["SLACK_CHANNEL_ID"]
debug_mode=os.environ.get("DEBUG_MODE","0")=="1"
run_mode=(os.environ.get("RUN_MODE","").strip().lower() or ("dev" if debug_mode else "prod"))
local_mode=run_mode=="local"
local_excel_path=(os.environ.get("LOCAL_EXCEL_PATH","./BlackBox_local.xlsx").strip() or "./BlackBox_local.xlsx")
client_id=os.environ.get("AZURE_CLIENT_ID","").strip()
refresh_token=os.environ.get("GRAPH_REFRESH_TOKEN","")
onedrive_upn=os.environ.get("ONEDRIVE_UPN","").strip()
requested_onedrive_file_path=(os.environ.get("ONEDRIVE_FILE_PATH","/Documents/BlackBox.xlsx").strip() or "/Documents/BlackBox.xlsx")
classification_guide_path=(os.environ.get("BLACKBOX_GUIDE_PATH","blackbox-categorias.md").strip() or "blackbox-categorias.md")
blackbox_guide_required=(os.environ.get("BLACKBOX_GUIDE_REQUIRED","1").strip()=="1")
classification_guide_max_chars=max(1200, int(os.environ.get("BLACKBOX_GUIDE_MAX_CHARS","6000")))
strict_md_classification=(os.environ.get("STRICT_MD_CLASSIFICATION","1").strip()=="1")
app_base_dir=os.path.dirname(os.path.abspath(__file__))

def build_dev_onedrive_path(path):
    p=(path or "").strip() or "/Documents/BlackBox.xlsx"
    if p.lower().endswith("_dev.xlsx"):
        return p
    if p.lower().endswith(".xlsx"):
        return p[:-5] + "_dev.xlsx"
    return p + "_dev.xlsx"

onedrive_file_path=build_dev_onedrive_path(requested_onedrive_file_path) if run_mode=="dev" else requested_onedrive_file_path
# target_hour_local eliminado - ya no se usa restricción de hora
dev_team_member_ids=[i.strip() for i in os.environ.get("DEV_TEAM_MEMBER_IDS","").split(",") if i.strip()]
refresh_token_path=os.environ.get("REFRESH_TOKEN_PATH","/data/graph_refresh_token")
device_flow_wait_seconds=int(os.environ.get("DEVICE_FLOW_WAIT_SECONDS","600"))  # 10 min
graph_scope=os.environ.get("GRAPH_SCOPE","offline_access Files.ReadWrite").strip() or "offline_access Files.ReadWrite"
groq_api_key=os.environ.get("GROQ_API_KEY","").strip()
groq_model=os.environ.get("GROQ_MODEL","qwen/qwen3-32b").strip() or "qwen/qwen3-32b"
gemini_api_key=os.environ.get("GEMINI_API_KEY","").strip()
gemini_model=os.environ.get("GEMINI_MODEL","gemini-2.5-flash").strip() or "gemini-2.5-flash"
ai_provider_preference=(os.environ.get("AI_PROVIDER","auto").strip().lower() or "auto")

expected_columns=[
    "Fecha aproximada",
    "SLACK",
    "Tipo de función",
    "Módulo funcional",
    "Causa raíz",
    "Comentarios",
    "Categoría Soporte (estandarizado para reportes)",
    "Propuesta (Tarea en ClickUp cuando sea desarrollable /Cambio sistema)",
    "ESTADO FINAL",
    "resumen ia"
]
legacy_diagnosis_column="Diagnóstico causa raíz"
hyperlink_formula_pattern=re.compile(r'^=HYPERLINK\("([^"]*)","((?:[^"]|"")*)"\)$')
blackbox_guide_cache=None

tipo_funcion_labels=[
    "Incidencia",
    "Duda",
    "Idea",
    "Soporte operativo",
    "Aviso",
]

modulo_funcional_labels=[
    "Contratos",
    "Cobros",
    "Pagos",
    "Liquidaciones",
    "Conciliación bancaria",
    "IPC / UF / Reajuste",
    "Propiedades / Unidades",
    "Stakeholders / Figuras",
    "Portal Arrendatarios",
    "CRM / Bot WhatsApp",
    "Corretaje",
    "Planner",
    "Servicios básicos",
    "Reportes / Exports",
    "Permisos / Auth",
    "Infraestructura",
]

causa_raiz_labels=[
    "Dato legacy / Dato histórico con problemas",
    "Lógica de negocio",
    "Caso borde no cubierto",
    "Gap de QA",
    "Integración externa",
    "UX / Capacitación",
    "Performance",
    "Configuración / Deploy",
    "Desconocido",
    "No aplica",
]

causa_raiz_requerida_para={"Incidencia","Soporte operativo"}

tipo_funcion_aliases={
    "incidente":"Incidencia",
    "incidencias":"Incidencia",
    "pregunta":"Duda",
    "consulta":"Duda",
    "dudas":"Duda",
    "idea de mejora":"Idea",
    "sugerencia":"Idea",
    "mejora":"Idea",
    "soporte":"Soporte operativo",
    "soporte manual":"Soporte operativo",
    "operativo":"Soporte operativo",
    "aviso informativo":"Aviso",
    "anuncio":"Aviso",
    "agradecimiento":"Aviso",
}

modulo_funcional_aliases={
    "conciliacion":"Conciliación bancaria",
    "portal":"Portal Arrendatarios",
    "crm":"CRM / Bot WhatsApp",
    "bot whatsapp":"CRM / Bot WhatsApp",
    "reportes":"Reportes / Exports",
    "exports":"Reportes / Exports",
    "auth":"Permisos / Auth",
    "permisos":"Permisos / Auth",
    "ipc":"IPC / UF / Reajuste",
    "uf":"IPC / UF / Reajuste",
}

causa_raiz_aliases={
    "dato legacy":"Dato legacy / Dato histórico con problemas",
    "dato historico con problemas":"Dato legacy / Dato histórico con problemas",
    "dato historico":"Dato legacy / Dato histórico con problemas",
    "logica negocio":"Lógica de negocio",
    "logica de negocio":"Lógica de negocio",
    "caso borde":"Caso borde no cubierto",
    "qa":"Gap de QA",
    "gap qa":"Gap de QA",
    "integracion":"Integración externa",
    "integracion externa":"Integración externa",
    "ux":"UX / Capacitación",
    "capacitacion":"UX / Capacitación",
    "performance":"Performance",
    "configuracion":"Configuración / Deploy",
    "deploy":"Configuración / Deploy",
    "desconocida":"Desconocido",
    "n a":"No aplica",
    "na":"No aplica",
    "vacio":"No aplica",
}
user_display_names={
    "U06BJ8JQ7B8":"agustin",
    "U07UBRSER6D":"neil",
    "U05D1H8JPEJ":"vico",
    "U06BF8NPZ5J":"luna"
}
mention_user_pattern=re.compile(r"<@([A-Z0-9]+)>")

def get_user_label(user_id):
    uid=(user_id or "").strip()
    if not uid:
        return "desconocido"
    display_name=user_display_names.get(uid)
    return display_name if display_name else uid

def replace_known_user_ids(text):
    out=str(text or "")
    def replace_mention(match):
        return get_user_label(match.group(1))
    out=mention_user_pattern.sub(replace_mention, out)
    return out

def canonicalize_known_user_labels(text):
    out=str(text or "")
    for user_id, display_name in user_display_names.items():
        canonical=display_name
        out=re.sub(
            rf"\b{re.escape(display_name)}\s*\(\s*{re.escape(user_id)}\s*\)",
            canonical,
            out,
            flags=re.IGNORECASE
        )
        out=re.sub(
            rf"\b{re.escape(user_id)}\s*\(\s*{re.escape(display_name)}\s*\)",
            canonical,
            out,
            flags=re.IGNORECASE
        )
        out=re.sub(
            rf"\b{re.escape(user_id)}\s*==\s*{re.escape(display_name)}\b",
            canonical,
            out,
            flags=re.IGNORECASE
        )
        out=re.sub(rf"\b{re.escape(user_id)}\b", canonical, out)
    return out

def now_scl():
    return datetime.now(tz=ZoneInfo("America/Santiago"))

# Función should_run() eliminada - ya no se usa restricción de hora

def load_refresh_token() -> str:
    """Lee refresh token desde archivo persistente (si existe) o desde ENV."""
    try:
        p=(refresh_token_path or "").strip()
        if p and os.path.exists(p):
            with open(p,"r",encoding="utf-8") as f:
                t=f.read().strip()
                if t:
                    return t
    except Exception as e:
        print(f"[WARN] No se pudo leer REFRESH_TOKEN_PATH={refresh_token_path}: {e}")
    return (os.environ.get("GRAPH_REFRESH_TOKEN","").strip() or "")

def save_refresh_token(token: str):
    """Guarda refresh token en archivo persistente para próximos runs."""
    if not token:
        return
    try:
        p=(refresh_token_path or "").strip()
        if not p:
            return
        d=os.path.dirname(p)
        if d:
            os.makedirs(d,exist_ok=True)
        with open(p,"w",encoding="utf-8") as f:
            f.write(token.strip())
        print(f"[INFO] Refresh token guardado en {p}")
    except Exception as e:
        print(f"[WARN] No se pudo guardar refresh token en {refresh_token_path}: {e}")

def acquire_token():
    tenant = os.environ.get("AZURE_TENANT", "consumers").strip() or "consumers"
    token_url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"

    if not client_id:
        raise RuntimeError("token: falta variable de entorno AZURE_CLIENT_ID")

    rt=load_refresh_token() or refresh_token

    def device_flow_token():
        device_url=f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/devicecode"
        dc=requests.post(device_url,data={"client_id":client_id,"scope":graph_scope},timeout=30)
        if dc.status_code>=400:
            raise RuntimeError(f"token: devicecode falló (status={dc.status_code}): {dc.text[:1000]}")
        flow=dc.json()
        msg=flow.get("message") or (
            f"Autorización requerida.\n1) Abrí {flow.get('verification_uri')}\n2) Ingresá el código: {flow.get('user_code')}"
        )
        print("[WARN] Se requiere re-login de Microsoft. Device Code Flow:")
        print(msg)
        # No enviar mensajes a Slack: solo log.

        interval=int(flow.get("interval",5))
        deadline=time.time()+min(int(flow.get("expires_in",900)),device_flow_wait_seconds)
        while time.time()<deadline:
            time.sleep(interval)
            tr=requests.post(
                token_url,
                data={
                    "grant_type":"urn:ietf:params:oauth:grant-type:device_code",
                    "client_id":client_id,
                    "device_code":flow.get("device_code"),
                },
                timeout=30
            )
            if tr.status_code==200:
                tok=tr.json()
                new_rt=tok.get("refresh_token")
                if new_rt:
                    save_refresh_token(new_rt)
                at=tok.get("access_token")
                if not at:
                    raise RuntimeError(f"token: device flow sin access_token: {str(tok)[:800]}")
                return at

            # Respuestas esperables mientras se autoriza
            try:
                err=tr.json()
            except ValueError:
                raise RuntimeError(f"token: device flow respuesta no-JSON (status={tr.status_code}): {tr.text[:500]}")
            code=err.get("error")
            if code in {"authorization_pending","slow_down"}:
                if code=="slow_down":
                    interval+=5
                continue
            raise RuntimeError(f"token: device flow falló: {err}")

        raise RuntimeError("token: expiró la espera de autorización (device flow). Reintentá luego de autorizar.")

    # Si no hay refresh token, necesitamos device flow sí o sí.
    if not rt:
        return device_flow_token()

    # Para refresh_token en v2.0, el parámetro scope es opcional; si se incluye, debe ser subconjunto del original.
    # Intentamos primero con scope (comportamiento actual) y, si falla con error de scope, reintentamos sin scope.
    base_data = {
        "client_id": client_id,
        "refresh_token": rt,
        "grant_type": "refresh_token",
    }
    attempts = [
        {**base_data, "scope": "offline_access Files.ReadWrite"},
        base_data,
    ]

    last_err = None
    for data in attempts:
        try:
            r = requests.post(token_url, data=data, timeout=30)
        except requests.RequestException as e:
            last_err = f"token: error de red llamando a {token_url}: {e}"
            continue

        if r.status_code < 400:
            try:
                payload = r.json()
            except ValueError:
                raise RuntimeError(f"token: respuesta no-JSON (status {r.status_code}) desde {token_url}: {r.text[:500]}")

            access_token = payload.get("access_token")
            if not access_token:
                raise RuntimeError(f"token: respuesta sin access_token desde {token_url}: {str(payload)[:800]}")
            return access_token

        # Error HTTP: extraer detalles (sin imprimir secretos)
        try:
            err = r.json()
        except ValueError:
            err = {"raw": r.text[:1000]}

        error_code = err.get("error") or "unknown_error"
        error_desc = err.get("error_description") or err.get("raw") or ""
        request_id = r.headers.get("request-id") or r.headers.get("x-ms-request-id") or r.headers.get("client-request-id") or ""

        last_err = (
            "token: fallo al refrescar access token "
            f"(tenant={tenant}, status={r.status_code}, error={error_code})"
            + (f", request_id={request_id}" if request_id else "")
            + (f": {error_desc}" if error_desc else "")
        )

        # Si el error sugiere un problema de scope, probamos el siguiente intento (sin scope).
        if error_code in {"invalid_scope"} or "scope" in str(error_desc).lower():
            continue
        # Si el refresh token quedó inválido (invalid_grant), caemos al Device Code Flow.
        # En cuentas personales esto puede ocurrir periódicamente y requiere re-login.
        if error_code=="invalid_grant":
            return device_flow_token()
        break

    # Fallback extra: si por algún motivo no matcheó antes, pero el último error fue invalid_grant, intentar device flow.
    if last_err and "error=invalid_grant" in last_err:
        return device_flow_token()

    raise RuntimeError(last_err or "token: fallo desconocido al refrescar access token")

def gget(url,token):
    r=requests.get(url,headers={"Authorization":f"Bearer {token}"})
    if r.status_code>=400:
        raise RuntimeError("get")
    return r

def gput(url,token,data,content_type):
    r=requests.put(url,headers={"Authorization":f"Bearer {token}","Content-Type":content_type},data=data)
    if r.status_code>=400:
        print(f"[ERROR] Error en PUT: {r.status_code} - {r.text}")
        # No lanzar excepción para errores 423 (archivo bloqueado) o 409 (conflicto)
        if r.status_code in [423, 409]:
            print(f"[WARN] Archivo bloqueado o en conflicto, continuando...")
            return r
        raise RuntimeError("put")
    return r

def ensure_file(token):
    meta=f"https://graph.microsoft.com/v1.0/users/{onedrive_upn}/drive/root:{onedrive_file_path}"
    r=requests.get(meta,headers={"Authorization":f"Bearer {token}"})
    if r.status_code==404:
        buf=io.BytesIO()
        wb=Workbook()
        ws=wb.active
        ws.title="TMP"
        wb.save(buf)
        buf.seek(0)
        upload=f"https://graph.microsoft.com/v1.0/users/{onedrive_upn}/drive/root:{onedrive_file_path}:/content"
        gput(upload,token,buf.getvalue(),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        print(f"[WARN] Archivo no existía, creado nuevo en {onedrive_file_path}")
    elif r.status_code>=400:
        raise RuntimeError("meta")

def ensure_local_file(path):
    p=(path or "").strip() or "./BlackBox_local.xlsx"
    if os.path.exists(p):
        return
    parent=os.path.dirname(p)
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb=Workbook()
    ws=wb.active
    ws.title="TMP"
    wb.save(p)
    print(f"[INFO] Archivo local creado: {os.path.abspath(p)}")

def dl_excel_local(path):
    with open(path, "rb") as f:
        return io.BytesIO(f.read())

def up_excel_local(path, bio):
    p=(path or "").strip() or "./BlackBox_local.xlsx"
    parent=os.path.dirname(p)
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(p, "wb") as f:
        f.write(bio.getvalue())
    print(f"[INFO] Excel guardado en local: {os.path.abspath(p)}")
    return True

def dl_excel(token):
    url=f"https://graph.microsoft.com/v1.0/users/{onedrive_upn}/drive/root:{onedrive_file_path}:/content"
    return io.BytesIO(gget(url,token).content)

def up_excel(token,bio):
    url=f"https://graph.microsoft.com/v1.0/users/{onedrive_upn}/drive/root:{onedrive_file_path}:/content"
    def save_local_backup():
        backup_filename=f"backup_blackbox_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            with open(backup_filename, "wb") as f:
                f.write(bio.getvalue())
            print(f"[INFO] Archivo guardado localmente como respaldo: {backup_filename}")
        except Exception as backup_error:
            print(f"[WARN] No se pudo guardar respaldo local: {backup_error}")

    try:
        response=gput(url,token,bio.getvalue(),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        status_code=getattr(response, "status_code", 500)
        if status_code < 400:
            print(f"[INFO] Archivo subido exitosamente a OneDrive")
            return True

        if status_code in [423, 409]:
            print("[WARN] No se actualizó el archivo principal por lock/conflicto (423/409)")
            try:
                timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
                base_path=onedrive_file_path.rsplit("/", 1)[0] if "/" in onedrive_file_path else ""
                base_name=onedrive_file_path.rsplit("/", 1)[-1].rsplit(".", 1)[0] if "." in onedrive_file_path else onedrive_file_path
                extension=onedrive_file_path.rsplit(".", 1)[-1] if "." in onedrive_file_path else "xlsx"

                backup_path=f"{base_path}/{base_name}_backup_{timestamp}.{extension}" if base_path else f"{base_name}_backup_{timestamp}.{extension}"
                backup_url=f"https://graph.microsoft.com/v1.0/users/{onedrive_upn}/drive/root:{backup_path}:/content"
                backup_response=gput(backup_url, token, bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                if getattr(backup_response, "status_code", 500) < 400:
                    print(f"[INFO] Archivo bloqueado - copia creada en OneDrive: {backup_path}")
                else:
                    print(f"[WARN] No se pudo crear copia en OneDrive (status={backup_response.status_code})")
            except Exception as backup_error:
                print(f"[ERROR] Error al crear copia en OneDrive: {backup_error}")

            save_local_backup()
            return False

        print(f"[ERROR] Error al subir archivo (status={status_code})")
        save_local_backup()
        print(f"[WARN] Continuando sin subir a OneDrive debido al error")
        return False
    except Exception as e:
        print(f"[ERROR] Error al subir archivo: {e}")
        save_local_backup()
        print(f"[WARN] Continuando sin subir a OneDrive debido al error")
        return False

def get_blackbox_guide_text():
    global blackbox_guide_cache
    if blackbox_guide_cache is not None:
        return blackbox_guide_cache
    default_text=(
        "Dimensión 1 (Tipo de entrada): Incidencia, Duda, Idea, Soporte operativo, Aviso.\n"
        "Dimensión 2 (Módulo funcional): Contratos, Cobros, Pagos, Liquidaciones, Conciliación bancaria, "
        "IPC / UF / Reajuste, Propiedades / Unidades, Stakeholders / Figuras, Portal Arrendatarios, "
        "CRM / Bot WhatsApp, Corretaje, Planner, Servicios básicos, Reportes / Exports, Permisos / Auth, Infraestructura.\n"
        "Dimensión 3 (Causa raíz): Dato legacy / Dato histórico con problemas, Lógica de negocio, Caso borde no cubierto, "
        "Gap de QA, Integración externa, UX / Capacitación, Performance, Configuración / Deploy, Desconocido. "
        "Para Duda, Idea y Aviso, usar No aplica."
    )
    try:
        guide_candidates=[classification_guide_path]
        if not os.path.isabs(classification_guide_path):
            guide_candidates.append(os.path.join(app_base_dir, classification_guide_path))
        resolved_path=None
        for candidate in guide_candidates:
            if candidate and os.path.exists(candidate):
                resolved_path=candidate
                break
        if not resolved_path:
            raise FileNotFoundError(
                f"No existe guía en ninguna ruta candidata: {guide_candidates}"
            )

        with open(resolved_path, "r", encoding="utf-8") as f:
            raw=f.read().strip()
            if not raw:
                if blackbox_guide_required:
                    raise RuntimeError(f"La guía {resolved_path} está vacía y es obligatoria")
                blackbox_guide_cache=default_text
            else:
                blackbox_guide_cache=raw
                print(f"[INFO] Guía BlackBox cargada desde {resolved_path} ({len(raw)} chars)")
    except Exception as e:
        if blackbox_guide_required:
            raise RuntimeError(
                f"No se pudo leer la guía BlackBox obligatoria en {classification_guide_path}: {e}"
            ) from e
        print(f"[WARN] No se pudo leer {classification_guide_path}: {e}")
        blackbox_guide_cache=default_text
    return blackbox_guide_cache

def get_blackbox_guide_excerpt():
    guide=get_blackbox_guide_text()
    if len(guide) <= classification_guide_max_chars:
        return guide
    reglas_marker="## Reglas de clasificación"
    idx=guide.find(reglas_marker)
    if idx == -1:
        return guide[:classification_guide_max_chars]
    tail_budget=max(800, classification_guide_max_chars//3)
    tail=guide[idx:idx+tail_budget]
    head_budget=max(400, classification_guide_max_chars - len(tail) - 8)
    head=guide[:head_budget]
    return f"{head}\n...\n{tail}"

def fetch_messages(oldest=None, latest=None):
    c=WebClient(token=slack_bot_token)
    out=[]
    cur=None
    while True:
        res=c.conversations_history(
            channel=channel_id,
            limit=1000,
            cursor=cur,
            oldest=oldest,
            latest=latest
        )
        out.extend(res.get("messages",[]))
        cur=res.get("response_metadata",{}).get("next_cursor")
        if not cur:
            break
    return out

def tz_dt(ts):
    return datetime.fromtimestamp(float(ts),tz=timezone.utc).astimezone(ZoneInfo("America/Santiago"))

def sanitize_text(text, max_len=None, escape_quotes=False):
    if text is None:
        return ""
    out=replace_known_user_ids(text).replace("\n"," ").replace("\r"," ")
    out=re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]"," ",out)
    out=re.sub(r"\s+"," ",out).strip()
    if max_len and len(out)>max_len:
        out=out[:max_len-3]+"..."
    if escape_quotes:
        out=out.replace('"','""')
    return out

def build_hyperlink_formula(url, label):
    safe_label=sanitize_text(label, max_len=200, escape_quotes=True)
    if not safe_label:
        safe_label="Abrir mensaje"
    formula=f'=HYPERLINK("{url}","{safe_label}")'
    if not hyperlink_formula_pattern.match(formula):
        formula=f'=HYPERLINK("{url}","Abrir mensaje")'
    return formula

def build_slack_hyperlink(ts, text):
    ts_formatted=str(ts or "").replace(".","")
    slack_link=f"https://mq-sede.slack.com/archives/{channel_id}/p{ts_formatted}"
    return build_hyperlink_formula(slack_link, text), slack_link

def fetch_thread_replies(client, thread_ts):
    if not thread_ts:
        return []
    out=[]
    cur=None
    while True:
        try:
            res=client.conversations_replies(
                channel=channel_id,
                ts=thread_ts,
                limit=200,
                cursor=cur
            )
        except SlackApiError as e:
            err=e.response.get("error") if getattr(e, "response", None) else str(e)
            print(f"[WARN] No se pudo leer hilo {thread_ts}: {err}")
            break
        out.extend(res.get("messages",[]))
        cur=res.get("response_metadata",{}).get("next_cursor")
        if not cur:
            break
    return out

def build_comments_from_thread(replies, root_ts):
    comentarios=[]
    for msg in replies:
        if msg.get("ts")==root_ts:
            continue
        txt=sanitize_text(msg.get("text",""), max_len=600)
        if not txt:
            continue
        user_id=msg.get("user")
        author=get_user_label(user_id) if user_id else (msg.get("bot_id") or "desconocido")
        ts=msg.get("ts")
        if ts:
            fecha=tz_dt(ts).strftime("%Y-%m-%d %H:%M:%S")
            comentarios.append(f"[{fecha}] {author}: {txt}")
        else:
            comentarios.append(f"{author}: {txt}")
    merged=" | ".join(comentarios).strip()
    return merged[:3997]+"..." if len(merged)>4000 else merged

def create_groq_client():
    if not groq_api_key:
        if not gemini_api_key:
            print("[WARN] GROQ_API_KEY no configurada y GEMINI_API_KEY ausente; se usará fallback.")
        return None
    if Groq is None:
        if not gemini_api_key:
            print("[WARN] Librería groq no instalada y GEMINI_API_KEY ausente; se usará fallback.")
        return None
    try:
        return Groq(api_key=groq_api_key)
    except Exception as e:
        print(f"[WARN] No se pudo crear cliente Groq: {e}")
        return None

def call_gemini_text(prompt, max_output_tokens=1024, temperature=0, response_mime_type=None):
    if not gemini_api_key:
        return ""
    url=f"https://generativelanguage.googleapis.com/v1beta/models/{gemini_model}:generateContent?key={gemini_api_key}"
    generation_config={
        "temperature": temperature,
        "maxOutputTokens": max(32, int(max_output_tokens)),
    }
    if response_mime_type:
        generation_config["responseMimeType"]=response_mime_type
    payload={
        "contents":[{"parts":[{"text": prompt}]}],
        "generationConfig": generation_config,
    }
    try:
        r=requests.post(url, json=payload, timeout=45)
        if r.status_code >= 400:
            print(f"[WARN] Gemini error HTTP {r.status_code}: {sanitize_text(r.text, max_len=300)}")
            return ""
        data=r.json()
        texts=[]
        for cand in data.get("candidates", []):
            content=(cand or {}).get("content") or {}
            for part in content.get("parts", []):
                txt=(part or {}).get("text")
                if txt:
                    texts.append(txt)
        return "\n".join(texts).strip()
    except Exception as e:
        print(f"[WARN] Error llamando Gemini: {e}")
        return ""

def call_groq_text(groq_client, prompt, max_completion_tokens=1024, temperature=0, top_p=1, stream=False):
    if groq_client is None:
        return ""
    try:
        completion=groq_client.chat.completions.create(
            model=groq_model,
            messages=[{"role":"user","content":prompt}],
            temperature=temperature,
            max_completion_tokens=max_completion_tokens,
            top_p=top_p,
            reasoning_effort="default",
            stream=stream,
            stop=None
        )
        if stream:
            output=[]
            for chunk in completion:
                output.append(chunk.choices[0].delta.content or "")
            return "".join(output).strip()
        return (completion.choices[0].message.content or "").strip()
    except Exception as e:
        print(f"[WARN] Error llamando Groq: {e}")
        return ""

def generate_llm_text(groq_client, prompt, max_completion_tokens=1024, temperature=0, top_p=1, stream=False, response_mime_type=None):
    provider=ai_provider_preference if ai_provider_preference in {"auto", "gemini", "groq"} else "auto"
    orders={
        "gemini": ("gemini", "groq"),
        "groq": ("groq", "gemini"),
        "auto": ("gemini", "groq") if gemini_api_key else ("groq", "gemini"),
    }
    for candidate in orders[provider]:
        if candidate=="gemini":
            out=call_gemini_text(
                prompt=prompt,
                max_output_tokens=max_completion_tokens,
                temperature=temperature,
                response_mime_type=response_mime_type
            )
        else:
            out=call_groq_text(
                groq_client=groq_client,
                prompt=prompt,
                max_completion_tokens=max_completion_tokens,
                temperature=temperature,
                top_p=top_p,
                stream=stream
            )
        if out:
            return out
    return ""

def generate_ai_summary(groq_client, main_text, comments_text, root_user_id=None):
    if not comments_text:
        return "Sin comentarios en el hilo. Conclusión: Sin información suficiente."
    if groq_client is None and not gemini_api_key:
        return "No se pudo generar resumen IA (sin cliente de IA). Conclusión: Sin información suficiente."

    root_author=get_user_label(root_user_id)
    prompt=(
        "Analiza este hilo de soporte de Slack y entrega un resumen breve.\n"
        "Reglas de autoría (obligatorias):\n"
        "- El autor del mensaje principal es exactamente el siguiente: "
        f"{root_author}\n"
        "- No confundas menciones con autoría. Un nombre/ID mencionado en el texto no implica que esa persona reportó.\n"
        "- Si mencionas personas y existe mapeo de ID->nombre, usa el nombre. Si no existe mapeo, usa el ID.\n"
        "- Si no hay certeza de autoría para una acción, usa redacción neutral ('se reporta', 'se comenta').\n\n"
        "Mensaje principal:\n"
        f"{sanitize_text(main_text, max_len=2000)}\n\n"
        "Comentarios del hilo:\n"
        f"{sanitize_text(comments_text, max_len=6000)}\n\n"
        "Responde en español con máximo 3 frases. "
        "No incluyas razonamiento interno, ni etiquetas XML/HTML como <think>. "
        "La última frase debe empezar exactamente con 'Conclusión:' y usar solo una etiqueta: "
        "'Resuelto', 'No resuelto' o 'Sin información suficiente'."
    )

    try:
        llm_output=generate_llm_text(
            groq_client=groq_client,
            prompt=prompt,
            temperature=0.6,
            max_completion_tokens=4096,
            top_p=0.95,
            stream=True
        )
        summary=normalize_ai_summary(llm_output)
        if summary:
            return summary[:4997]+"..." if len(summary)>5000 else summary
    except Exception as e:
        print(f"[WARN] Error generando resumen IA: {e}")

    return "No se pudo generar resumen IA. Conclusión: Sin información suficiente."

def normalize_ai_summary(text):
    s=(text or "").strip()
    s=re.sub(r"<think>.*?</think>\s*","",s,flags=re.IGNORECASE|re.DOTALL)
    s=s.replace("**","").replace("__","")
    s=canonicalize_known_user_labels(s)
    s=sanitize_text(s)
    s=canonicalize_known_user_labels(s)
    return s

def normalize_ai_status_output(text):
    s=(text or "").strip()
    if not s:
        return ""
    s=s.replace("**","").replace("__","")
    normalized=normalize_for_status_matching(s)
    if not normalized:
        return ""
    if "idea" in normalized:
        return "IDEA"
    if "anuncio" in normalized or "agradecimiento" in normalized:
        return "Anuncio/Agradecimiento"
    empty_signals=("vacio", "sin estado", "sin categoria", "ninguno", "ninguna", "no aplica")
    if any(sig in normalized for sig in empty_signals):
        return ""
    return ""

def generate_ai_status(groq_client, main_text, comments_text):
    fallback_status=infer_auto_status(main_text)
    if is_incident_like(f"{main_text or ''} {comments_text or ''}"):
        return ""
    if groq_client is None and not gemini_api_key:
        return fallback_status

    prompt=(
        "Clasifica el mensaje de Slack para la columna 'ESTADO FINAL'.\n"
        "Responde SOLO con una etiqueta exacta (sin explicación):\n"
        "IDEA\n"
        "Anuncio/Agradecimiento\n"
        "VACIO\n\n"
        "Reglas:\n"
        "- IDEA: solicitud o sugerencia de nueva funcionalidad/cambio.\n"
        "- Anuncio/Agradecimiento: comunicación de novedades/despliegues o agradecimientos.\n"
        "- VACIO: cuando no aplica ninguna etiqueta.\n\n"
        "Mensaje principal:\n"
        f"{sanitize_text(main_text, max_len=2500)}\n\n"
        "Comentarios del hilo (puede estar vacío):\n"
        f"{sanitize_text(comments_text, max_len=3500)}\n"
    )

    try:
        llm_output=generate_llm_text(
            groq_client=groq_client,
            prompt=prompt,
            temperature=0,
            max_completion_tokens=64,
            top_p=1,
            stream=False
        )
        ai_status=normalize_ai_status_output(llm_output)
        if is_incident_like(f"{main_text or ''} {comments_text or ''}"):
            return ""
        if ai_status in {"IDEA", "Anuncio/Agradecimiento", ""}:
            return ai_status
    except Exception as e:
        print(f"[WARN] Error generando estado IA: {e}")

    return fallback_status

def normalize_for_status_matching(text):
    s=sanitize_text(text).lower()
    replacements=str.maketrans("áéíóúüñ","aeiouun")
    s=s.translate(replacements)
    s=re.sub(r"[^a-z0-9 ]+"," ",s)
    s=re.sub(r"\s+"," ",s).strip()
    return s

def resolve_allowed_label(raw_value, allowed_labels, aliases=None, default=""):
    normalized=normalize_for_status_matching(raw_value)
    if not normalized:
        return default
    label_lookup={normalize_for_status_matching(label): label for label in allowed_labels}
    if normalized in label_lookup:
        return label_lookup[normalized]

    compact=normalized.replace(" ","")
    for normalized_label, original in label_lookup.items():
        if compact == normalized_label.replace(" ",""):
            return original

    for alias, target in (aliases or {}).items():
        alias_norm=normalize_for_status_matching(alias)
        if normalized == alias_norm or compact == alias_norm.replace(" ",""):
            return target

    for normalized_label, original in label_lookup.items():
        if len(normalized)>=5 and (normalized in normalized_label or normalized_label in normalized):
            return original

    return default

def is_incident_like(text):
    s=normalize_for_status_matching(text)
    if not s:
        return False
    incident_signals=(
        "error", "falla", "fallo", "problema", "incidencia", "caido", "caida",
        "no funciona", "no carga", "no responde",
        "no se genero", "no se generaron", "no se creo", "no se crearon",
        "no se aplico", "no cambio", "incorrecto", "duplicado",
        "urgente", "por favor revisar",
    )
    return any(sig in s for sig in incident_signals)

def final_status_for_tipo(tipo_funcion):
    if tipo_funcion == "Idea":
        return "IDEA"
    if tipo_funcion == "Aviso":
        return "Anuncio/Agradecimiento"
    return ""

def infer_tipo_funcion(main_text, comments_text):
    main=str(main_text or "")
    merged=normalize_for_status_matching(f"{main_text or ''} {comments_text or ''}")
    if not merged:
        return "Incidencia"
    if is_announcement_or_gratitude(main):
        return "Aviso"
    if is_idea_request(main):
        return "Idea"

    support_signals=(
        "necesito que", "manual", "backend", "query", "script", "reactivar",
        "cambiar estado", "borrar", "eliminar", "corregir dato", "ajustar dato",
    )
    if any(sig in merged for sig in support_signals):
        return "Soporte operativo"

    if is_incident_like(merged):
        return "Incidencia"

    doubt_signals=(
        "como", "donde", "quien", "que significa", "no entiendo",
        "se puede", "puedo", "por que", "porque",
    )
    if ("?" in main and not is_incident_like(merged)) or any(sig in merged for sig in doubt_signals):
        return "Duda"
    return "Incidencia"

def infer_modulo_funcional(main_text, comments_text):
    s=normalize_for_status_matching(f"{main_text or ''} {comments_text or ''}")
    if any(sig in s for sig in ("cobro", "cobros", "cuota", "cuotas", "ggcc")):
        return "Cobros"
    mapping=[
        ("Conciliación bancaria", ("conciliacion", "reconcili", "movimiento banc", "clearing", "cartola")),
        ("IPC / UF / Reajuste", ("ipc", "uf", "reajuste", "banco central", "clf")),
        ("Liquidaciones", ("liquidacion", "comision", "saldo neto")),
        ("Pagos", ("pago", "paymentintent", "fintoc", "nomina", "transferencia")),
        ("Cobros", ("cobro", "cuota", "multa", "deuda", "ggcc")),
        ("Contratos", ("contrato", "anexo", "renovacion", "garantia", "reserva")),
        ("Propiedades / Unidades", ("propiedad", "unidad", "departamento", "amenidad", "tipologia")),
        ("Stakeholders / Figuras", ("stakeholder", "owner", "person", "company", "figura", "propietario")),
        ("Portal Arrendatarios", ("portal", "arrendatario", "co deudor", "debtnotification", "link de pago")),
        ("CRM / Bot WhatsApp", ("whatsapp", "lead", "funnel", "conversation", "bot")),
        ("Corretaje", ("corretaje", "broker", "keybox", "mercadolibre", "prospect", "visita")),
        ("Planner", ("planner", "kanban", "notebook", "card", "tarjeta")),
        ("Servicios básicos", ("agua andina", "enel", "servicios basicos", "servicio basico")),
        ("Reportes / Exports", ("reporte", "excel", "zip", "pdf", "rentroll", "export")),
        ("Permisos / Auth", ("permiso", "rol", "auth", "login", "contrasena", "password")),
        ("Infraestructura", ("infraestructura", "worker", "celery", "s3", "sqs", "vercel", "timeout", "deploy")),
    ]
    for label, signals in mapping:
        if any(sig in s for sig in signals):
            return label
    return "Infraestructura"

def infer_causa_raiz(tipo_funcion, main_text, comments_text):
    if tipo_funcion not in causa_raiz_requerida_para:
        return "No aplica"
    s=normalize_for_status_matching(f"{main_text or ''} {comments_text or ''}")
    if not s:
        return "Desconocido"
    if any(sig in s for sig in ("legacy", "historico", "migracion", "dato antiguo", "version anterior")):
        return "Dato legacy / Dato histórico con problemas"
    if any(sig in s for sig in ("fintoc", "mercadolibre", "microsoft graph", "api banco central", "agua andina", "enel", "tuya")):
        return "Integración externa"
    if any(sig in s for sig in ("timeout", "504", "lento", "degradacion", "saturacion", "performance")):
        return "Performance"
    if any(sig in s for sig in ("deploy", "configuracion", "entorno", "produccion", "variable de entorno")):
        return "Configuración / Deploy"
    if any(sig in s for sig in ("no se probo", "no probado", "falta test", "qa", "regresion")):
        return "Gap de QA"
    if any(sig in s for sig in ("no entiende", "confusion", "capacitacion", "usabilidad", "ux")):
        return "UX / Capacitación"
    if any(sig in s for sig in ("caso borde", "no contemplado", "escenario especial")):
        return "Caso borde no cubierto"
    if any(sig in s for sig in ("regla", "calculo", "logica", "estado incorrecto", "duplicado")):
        return "Lógica de negocio"
    if any(sig in s for sig in ("no reproducible", "sin contexto", "falta informacion", "ambiguo")):
        return "Desconocido"
    return "Desconocido"

def infer_blackbox_classification(main_text, comments_text):
    tipo=infer_tipo_funcion(main_text, comments_text)
    modulo=infer_modulo_funcional(main_text, comments_text)
    causa=infer_causa_raiz(tipo, main_text, comments_text)
    return {
        "tipo_funcion": tipo,
        "modulo_funcional": modulo,
        "causa_raiz": causa,
    }

def parse_blackbox_classification_output(text):
    payload={}
    raw=(text or "").strip()
    if not raw:
        return payload

    candidate_blocks=[raw]
    if "```" in raw:
        candidate_blocks.extend(re.findall(r"```(?:json)?\s*(.*?)```", raw, flags=re.IGNORECASE|re.DOTALL))
    json_like=re.search(r"\{.*\}", raw, flags=re.DOTALL)
    if json_like:
        candidate_blocks.append(json_like.group(0))

    for candidate in candidate_blocks:
        content=(candidate or "").strip()
        if not content:
            continue
        try:
            obj=json.loads(content)
            if isinstance(obj, dict):
                payload=obj
                break
        except Exception:
            continue

    if payload:
        return payload

    line_patterns={
        "tipo_funcion": r"(?:tipo[_\s]*funcion|tipo[_\s]*entrada)\s*[:=]\s*(.+)",
        "modulo_funcional": r"(?:modulo[_\s]*funcional)\s*[:=]\s*(.+)",
        "causa_raiz": r"(?:causa[_\s]*raiz)\s*[:=]\s*(.+)",
    }
    for key, pattern in line_patterns.items():
        m=re.search(pattern, raw, flags=re.IGNORECASE)
        if m:
            payload[key]=sanitize_text(m.group(1), max_len=200)
    return payload

def generate_blackbox_classification(groq_client, main_text, comments_text):
    fallback=infer_blackbox_classification(main_text, comments_text)
    if groq_client is None and not gemini_api_key:
        if strict_md_classification:
            print("[WARN] Modo estricto .md: sin cliente de IA, usando clasificador determinístico basado en guía.")
        return fallback

    guide_text=get_blackbox_guide_excerpt()
    def build_prompt(guide_excerpt):
        return (
            "Usa la guía de clasificación del BlackBox y clasifica el mensaje en 3 dimensiones.\n"
            "Responde SOLO JSON válido (sin markdown) con este formato exacto:\n"
            "{\"tipo_funcion\":\"...\",\"modulo_funcional\":\"...\",\"causa_raiz\":\"...\"}\n\n"
            "Etiquetas permitidas para tipo_funcion:\n"
            + "\n".join(f"- {x}" for x in tipo_funcion_labels) + "\n\n"
            "Etiquetas permitidas para modulo_funcional:\n"
            + "\n".join(f"- {x}" for x in modulo_funcional_labels) + "\n\n"
            "Etiquetas permitidas para causa_raiz:\n"
            + "\n".join(f"- {x}" for x in causa_raiz_labels) + "\n\n"
            "Reglas obligatorias:\n"
            "- Si tipo_funcion es Duda, Idea o Aviso, causa_raiz debe ser 'No aplica'.\n"
            "- Si tipo_funcion es Incidencia o Soporte operativo y no hay evidencia suficiente, usar 'Desconocido'.\n"
            "- Elegir siempre una sola etiqueta por dimensión.\n\n"
            "Heurísticas clave para evitar errores:\n"
            "- Mensajes informativos de despliegue/automatización (ej: 'se creó alerta automática', 'ya quedó habilitado') => tipo_funcion='Aviso'.\n"
            "- Mensajes de falla operativa (ej: 'no se generaron cobros', 'da error', 'no funciona') => tipo_funcion='Incidencia'.\n"
            "- Si el texto menciona 'cobro/cobros/cuotas/GGCC', priorizar modulo_funcional='Cobros'.\n\n"
            "Guía base:\n"
            f"{guide_excerpt}\n\n"
            "Mensaje principal:\n"
            f"{sanitize_text(main_text, max_len=1800)}\n\n"
            "Comentarios del hilo:\n"
            f"{sanitize_text(comments_text, max_len=2500)}\n"
        )

    def request_payload(prompt_text):
        content=generate_llm_text(
            groq_client=groq_client,
            prompt=prompt_text,
            temperature=0,
            max_completion_tokens=256,
            top_p=1,
            stream=False,
            response_mime_type="application/json"
        )
        return parse_blackbox_classification_output(content), content

    try:
        payload={}
        raw_output=""
        prompt_variants=[
            build_prompt(guide_text),
            build_prompt(guide_text[:1400]),
        ]
        for prompt in prompt_variants:
            payload, raw_output=request_payload(prompt)
            if payload:
                break

        tipo=resolve_allowed_label(payload.get("tipo_funcion"), tipo_funcion_labels, tipo_funcion_aliases, default="")
        modulo=resolve_allowed_label(payload.get("modulo_funcional"), modulo_funcional_labels, modulo_funcional_aliases, default="")
        causa=resolve_allowed_label(payload.get("causa_raiz"), causa_raiz_labels, causa_raiz_aliases, default="")

        if strict_md_classification and (not tipo or not modulo or not causa):
            print(
                "[WARN] Modo estricto .md: salida IA inválida, "
                f"usando clasificador determinístico. payload={payload}, raw={sanitize_text(raw_output, max_len=220)}"
            )
            return fallback

        result={
            "tipo_funcion": tipo or fallback["tipo_funcion"],
            "modulo_funcional": modulo or fallback["modulo_funcional"],
            "causa_raiz": causa or fallback["causa_raiz"],
        }

        merged_text=f"{main_text or ''} {comments_text or ''}"
        normalized_merged=normalize_for_status_matching(merged_text)
        if is_announcement_or_gratitude(main_text):
            result["tipo_funcion"]="Aviso"
        elif is_incident_like(merged_text):
            result["tipo_funcion"]="Incidencia"
        if any(sig in normalized_merged for sig in ("cobro", "cobros", "cuota", "cuotas", "ggcc")):
            result["modulo_funcional"]="Cobros"

        if result["tipo_funcion"] not in causa_raiz_requerida_para:
            result["causa_raiz"]="No aplica"
        elif result["causa_raiz"]=="No aplica":
            result["causa_raiz"]=fallback["causa_raiz"] or "Desconocido"
        return result
    except Exception as e:
        if strict_md_classification:
            print(
                "[WARN] Modo estricto .md: error de IA, "
                f"usando clasificador determinístico. error={e}"
            )
            return fallback
        print(f"[WARN] Error clasificando BlackBox con IA: {e}")
        return fallback

def is_idea_request(text):
    s=normalize_for_status_matching(text)
    if not s:
        return False
    if "lista de deseos" in s or "wishlist" in s:
        return True
    request_intents=("porfa", "podrian", "pueden", "me gustaria", "sugerencia", "idea", "deberian")
    feature_actions=("agregar", "incluir", "sumar", "habilitar", "implementar", "filtro")
    return any(i in s for i in request_intents) and any(a in s for a in feature_actions)

def is_announcement_or_gratitude(text):
    s=normalize_for_status_matching(text)
    if not s:
        return False
    if is_incident_like(s):
        return False
    announcement_signals=(
        "pasando a produccion",
        "a produccion",
        "novedades",
        "nuevo en produccion",
        "nuevos elementos",
        "se agrego a produccion",
        "ya esta en produccion",
        "ya esta disponible",
        "ya quedo",
        "se creo alerta",
        "se creo automatica",
        "se creo automatizacion",
        "se habilito",
        "se implemento",
        "deploy",
        "despliegue",
        "gracias por su paciencia",
    )
    if any(sig in s for sig in announcement_signals):
        return True

    # Detecta anuncios del tipo "nuevo/nueva + funcionalidad" aunque no mencionen "producción".
    announcement_change_signals=(
        "nuevo", "nueva", "nuevos", "nuevas",
        "se agrego", "agregamos",
        "se habilito", "habilitamos",
        "se actualizo", "actualizamos",
        "mejora", "mejoras",
    )
    announcement_feature_signals=(
        "filtro", "funcionalidad", "opcion", "campo", "modulo", "reporte",
        "pantalla", "proceso", "flujo", "regla", "validacion", "credito", "creditos",
    )
    issue_signals=("error", "problema", "no funciona", "no me", "ayuda", "incidencia")
    has_announcement_tone=any(sig in s for sig in announcement_change_signals)
    has_feature_context=any(sig in s for sig in announcement_feature_signals)
    has_issue_tone=any(sig in s for sig in issue_signals)
    if has_announcement_tone and has_feature_context and not has_issue_tone:
        return True

    gratitude_signals=("gracias", "muchas gracias", "agradecido", "agradecida", "agradecimiento")
    return (
        any(sig in s for sig in gratitude_signals)
        and len(s.split()) <= 14
        and not any(sig in s for sig in issue_signals)
    )

def infer_auto_status(main_text):
    if is_incident_like(main_text):
        return ""
    if is_idea_request(main_text):
        return "IDEA"
    if is_announcement_or_gratitude(main_text):
        return "Anuncio/Agradecimiento"
    return ""

def build_df(msgs, existing_keys=None):
    datos=[]
    known_keys=set(existing_keys or [])
    slack_client=WebClient(token=slack_bot_token)
    groq_client=create_groq_client()

    for m in reversed(msgs):
        uid=m.get("user")
        if not uid:
            continue

        ts=m.get("ts")
        if not ts:
            continue

        slack_content, slack_link=build_slack_hyperlink(ts, m.get("text",""))
        key=slack_link or slack_content
        if key in known_keys:
            continue

        dt=tz_dt(ts)

        comments_for_ai=""
        reply_count=int(m.get("reply_count",0) or 0)
        if reply_count>0:
            thread_ts=m.get("thread_ts") or ts
            replies=fetch_thread_replies(slack_client, thread_ts)
            comments_for_ai=build_comments_from_thread(replies, ts)

        main_text=m.get("text","")
        resumen_ia=generate_ai_summary(groq_client, main_text, comments_for_ai, root_user_id=uid)
        blackbox_classification=generate_blackbox_classification(groq_client, main_text, comments_for_ai)
        auto_status=final_status_for_tipo(blackbox_classification["tipo_funcion"])

        datos.append({
            "Fecha aproximada":dt.strftime("%Y-%m-%d %H:%M:%S"),
            "SLACK":slack_content,
            "Tipo de función":blackbox_classification["tipo_funcion"],
            "Módulo funcional":blackbox_classification["modulo_funcional"],
            "Causa raíz":blackbox_classification["causa_raiz"],
            "Comentarios":"",
            "Categoría Soporte (estandarizado para reportes)":"",
            "Propuesta (Tarea en ClickUp cuando sea desarrollable /Cambio sistema)":"",
            "ESTADO FINAL":auto_status,
            "resumen ia":resumen_ia
        })
        known_keys.add(key)

    return pd.DataFrame(datos,columns=expected_columns) if datos else pd.DataFrame(columns=expected_columns)

def extract_hyperlink_url(cell_value):
    """
    Extrae la URL de una fórmula de Excel del tipo:
    =HYPERLINK("url","texto")
    Si no aplica o falla, retorna None.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    if not s.startswith("=HYPERLINK("):
        return None
    # Captura la primera cadena entre comillas (la URL)
    m = re.match(r'^=HYPERLINK\("([^"]+)"\s*,', s)
    return m.group(1) if m else None

def get_month_name_from_period(df):
    """Obtiene el nombre del mes del primer día del período"""
    if df.empty:
        return "Datos"
    first_date = df.iloc[0]["Fecha aproximada"]
    dt = datetime.strptime(first_date, "%Y-%m-%d %H:%M:%S")
    month_names = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                   "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    return month_names[dt.month - 1]

def apply_table_style(ws, num_rows):
    """Aplica estilo profesional a la tabla"""
    try:
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        highlight_blue_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        
        # Borde para todas las celdas
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Aplicar estilo al header (fila 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dejar fijo el header (fila 1)
        ws.freeze_panes = "A2"

        if num_rows <= 1:  # Solo header o sin datos
            return
        
        # Aplicar estilo a las filas de datos
        estado_col_idx = get_column_index(ws, "ESTADO FINAL")
        for row in range(2, min(num_rows + 1, ws.max_row + 1)):
            highlight_blue = False
            if estado_col_idx is not None:
                raw_status = ws.cell(row=row, column=estado_col_idx).value
                normalized_status = normalize_for_status_matching(raw_status)
                highlight_blue = any(
                    kw in normalized_status for kw in ("idea", "anuncio", "agradecimiento")
                )

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                if highlight_blue:
                    cell.fill = highlight_blue_fill
        
        # Ajustar ancho de columnas por nombre (robusto si el orden cambia)
        column_widths_by_name = {
            "Fecha aproximada": 23,
            "SLACK": 60,
            "Tipo de función": 20,
            "Módulo funcional": 26,
            "Causa raíz": 30,
            "Comentarios": 30,
            "Categoría Soporte (estandarizado para reportes)": 42,
            "Propuesta (Tarea en ClickUp cuando sea desarrollable /Cambio sistema)": 48,
            "ESTADO FINAL": 22,
            "resumen ia": 70
        }
        headers=normalize_header_row(ws)
        for idx, header in enumerate(headers, start=1):
            width=column_widths_by_name.get(header)
            if width is not None:
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width
        
        # Ajustar altura de filas
        max_rows = min(num_rows, 200)
        for row in range(1, max_rows + 1):
            ws.row_dimensions[row].height = 42
        
        # Asegurar que el zoom esté al 100%
        ws.sheet_view.zoomScale = 100
        
        print(f"[INFO] Estilo aplicado: {max_rows} filas, columnas ajustadas")
            
    except Exception as e:
        print(f"[WARN] Error aplicando estilo: {e}")
        # Continuar sin estilo si hay problemas

def collect_existing_slack_keys(wb):
    keys=set()
    for sheet_name in wb.sheetnames:
        ws=wb[sheet_name]
        if ws.max_row<=1:
            continue
        slack_col=get_column_index(ws, "SLACK", fallback=2)
        if slack_col is None:
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if len(row)>=slack_col and row[slack_col-1]:
                slack_content=str(row[slack_col-1]).strip()
                url=extract_hyperlink_url(slack_content)
                keys.add(url or slack_content)
    return keys

def collect_existing_row_locations(wb):
    locations={}
    for sheet_name in wb.sheetnames:
        ws=wb[sheet_name]
        if ws.max_row<=1:
            continue
        slack_col=get_column_index(ws, "SLACK", fallback=2)
        if slack_col is None:
            continue
        for row_idx in range(2, ws.max_row+1):
            slack_value=ws.cell(row=row_idx, column=slack_col).value
            if not slack_value:
                continue
            slack_content=str(slack_value).strip()
            key=extract_hyperlink_url(slack_content) or slack_content
            if key and key not in locations:
                locations[key]=(sheet_name, row_idx)
    return locations

def backfill_ai_for_existing_rows(wb, msgs, existing_row_locations):
    if not msgs or not existing_row_locations:
        return 0, 0, 0, 0

    slack_client=WebClient(token=slack_bot_token)
    groq_client=create_groq_client()
    reclassify_existing=(os.environ.get("RECLASSIFY_EXISTING", "0").strip()=="1")
    clear_non_ai_columns=(os.environ.get("CLEAR_NON_AI_COLUMNS", "0").strip()=="1")
    force_reclassify_misclassified=(os.environ.get("FORCE_RECLASSIFY_MISCLASSIFIED", "1").strip()=="1")
    checked=0
    summary_updated=0
    status_updated=0
    classification_updated=0

    for m in reversed(msgs):
        uid=m.get("user")
        ts=m.get("ts")
        if not uid or not ts:
            continue

        _, slack_link=build_slack_hyperlink(ts, m.get("text",""))
        key=slack_link
        if key not in existing_row_locations:
            continue

        sheet_name, row_idx = existing_row_locations[key]
        ws=wb[sheet_name]
        status_col=get_column_index(ws, "ESTADO FINAL", fallback=8)
        summary_col=get_column_index(ws, "resumen ia", fallback=9)
        if summary_col is None:
            continue
        tipo_col=get_column_index(ws, "Tipo de función")
        modulo_col=get_column_index(ws, "Módulo funcional")
        causa_col=get_column_index(ws, "Causa raíz")
        comentarios_col=get_column_index(ws, "Comentarios")
        categoria_col=get_column_index(ws, "Categoría Soporte (estandarizado para reportes)")

        status_cell=ws.cell(row=row_idx, column=status_col) if status_col is not None else None
        summary_cell=ws.cell(row=row_idx, column=summary_col)
        tipo_cell=ws.cell(row=row_idx, column=tipo_col) if tipo_col is not None else None
        modulo_cell=ws.cell(row=row_idx, column=modulo_col) if modulo_col is not None else None
        causa_cell=ws.cell(row=row_idx, column=causa_col) if causa_col is not None else None
        comentarios_cell=ws.cell(row=row_idx, column=comentarios_col) if comentarios_col is not None else None
        categoria_cell=ws.cell(row=row_idx, column=categoria_col) if categoria_col is not None else None

        has_status=bool(str(status_cell.value or "").strip()) if status_cell is not None else False
        has_summary=bool(str(summary_cell.value or "").strip())
        has_tipo=bool(str(tipo_cell.value or "").strip()) if tipo_cell is not None else False
        has_modulo=bool(str(modulo_cell.value or "").strip()) if modulo_cell is not None else False
        has_causa=bool(str(causa_cell.value or "").strip()) if causa_cell is not None else False

        main_text=m.get("text","")
        current_tipo=str(tipo_cell.value or "").strip() if tipo_cell is not None else ""
        auto_status=final_status_for_tipo(current_tipo)
        needs_status_update=(not has_status) and bool(auto_status)
        needs_summary_update=(not has_summary)
        existing_status=str(status_cell.value or "").strip() if status_cell is not None else ""
        incident_like_main=is_incident_like(main_text)
        force_reclassify=(
            force_reclassify_misclassified
            and incident_like_main
            and (
                current_tipo=="Idea"
                or existing_status=="IDEA"
            )
        )
        needs_classification_update=any(
            (
                not has_tipo,
                not has_modulo,
                not has_causa,
            )
        )
        if reclassify_existing:
            needs_classification_update=True
        if force_reclassify:
            needs_classification_update=True
            print(f"[INFO] Reclasificación forzada por señales de incidencia (ts={ts})")

        if needs_status_update and status_cell is not None:
            status_cell.value=auto_status
            has_status=True
            status_updated+=1

        if not needs_summary_update and not needs_classification_update:
            continue

        comments_for_ai=""
        reply_count=int(m.get("reply_count",0) or 0)
        if reply_count>0:
            thread_ts=m.get("thread_ts") or ts
            replies=fetch_thread_replies(slack_client, thread_ts)
            comments_for_ai=build_comments_from_thread(replies, ts)

        if needs_summary_update:
            new_summary=normalize_ai_summary(
                generate_ai_summary(groq_client, main_text, comments_for_ai, root_user_id=uid)
            )
            checked+=1
            if new_summary and new_summary != str(summary_cell.value or "").strip():
                summary_cell.value=new_summary
                summary_updated+=1

        if needs_classification_update:
            classification=generate_blackbox_classification(groq_client, main_text, comments_for_ai)
            updated_current_row=False
            if tipo_cell is not None and ((not has_tipo) or str(tipo_cell.value or "").strip() != classification["tipo_funcion"]):
                tipo_cell.value=classification["tipo_funcion"]
                updated_current_row=True
            if modulo_cell is not None and ((not has_modulo) or str(modulo_cell.value or "").strip() != classification["modulo_funcional"]):
                modulo_cell.value=classification["modulo_funcional"]
                updated_current_row=True
            if causa_cell is not None and ((not has_causa) or str(causa_cell.value or "").strip() != classification["causa_raiz"]):
                causa_cell.value=classification["causa_raiz"]
                updated_current_row=True
            if status_cell is not None:
                status_for_tipo=final_status_for_tipo(classification["tipo_funcion"])
                current_status=str(status_cell.value or "").strip()
                if current_status != status_for_tipo:
                    status_cell.value=status_for_tipo
                    status_updated+=1
            if clear_non_ai_columns:
                for cell in (comentarios_cell, categoria_cell):
                    if cell is not None and str(cell.value or "").strip() != "":
                        cell.value=""
                        updated_current_row=True
            if updated_current_row:
                classification_updated+=1

    if checked or status_updated or classification_updated:
        print(
            "[INFO] Backfill IA/estado/clasificación sobre filas existentes: "
            f"revisadas={checked}, resumen_actualizado={summary_updated}, "
            f"estado_actualizado={status_updated}, clasificacion_actualizada={classification_updated}"
        )
    return checked, summary_updated, status_updated, classification_updated

def normalize_header_row(ws):
    return [str(ws.cell(row=1, column=idx).value).strip() if ws.cell(row=1, column=idx).value is not None else "" for idx in range(1, ws.max_column + 1)]

def get_column_index(ws, column_name, fallback=None):
    headers=normalize_header_row(ws)
    for idx, name in enumerate(headers, start=1):
        if name == column_name:
            return idx
    return fallback

def ensure_column_exists(ws, column_name):
    idx=get_column_index(ws, column_name)
    if idx is not None:
        return idx, False
    new_idx=ws.max_column + 1
    ws.cell(row=1, column=new_idx, value=column_name)
    return new_idx, True

def build_row_values_for_sheet(ws, row_dict):
    headers=normalize_header_row(ws)
    return [row_dict.get(h, "") for h in headers]

def migrate_sheet_headers(ws):
    # Migración no destructiva: nunca borra filas/celdas existentes.
    added_any=False

    for col_name in expected_columns:
        _, added = ensure_column_exists(ws, col_name)
        if added:
            added_any=True

    legacy_idx=get_column_index(ws, legacy_diagnosis_column)
    causa_idx=get_column_index(ws, "Causa raíz")
    comentarios_idx=get_column_index(ws, "Comentarios")
    if legacy_idx is not None and ws.max_row > 1 and (causa_idx is not None or comentarios_idx is not None):
        for row_idx in range(2, ws.max_row + 1):
            legacy_val=ws.cell(row=row_idx, column=legacy_idx).value
            if legacy_val in (None, ""):
                continue
            if causa_idx is not None:
                current_causa=ws.cell(row=row_idx, column=causa_idx).value
                if current_causa is None or str(current_causa).strip()=="":
                    ws.cell(row=row_idx, column=causa_idx, value=legacy_val)
            if comentarios_idx is not None:
                current_comentario=ws.cell(row=row_idx, column=comentarios_idx).value
                if current_comentario is None or str(current_comentario).strip()=="":
                    ws.cell(row=row_idx, column=comentarios_idx, value=legacy_val)

    return added_any

def repair_invalid_hyperlinks_in_sheet(ws):
    repaired=0
    if ws.max_row<=1:
        return repaired
    slack_col=get_column_index(ws, "SLACK", fallback=2)
    if slack_col is None:
        return repaired
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=slack_col, max_col=slack_col):
        cell=row[0]
        val=cell.value
        if not isinstance(val,str) or not val.startswith("=HYPERLINK("):
            continue
        if hyperlink_formula_pattern.match(val):
            continue
        url=extract_hyperlink_url(val)
        if url:
            cell.value=build_hyperlink_formula(url, "Abrir mensaje")
            repaired+=1
    return repaired

def repair_ai_columns_in_sheet(ws):
    cleaned=0
    if ws.max_row<=1:
        return cleaned
    summary_col=get_column_index(ws, "resumen ia", fallback=9)
    if summary_col is None:
        return cleaned
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        summary_cell=row[summary_col-1] if len(row)>=summary_col else None
        if summary_cell is not None and isinstance(summary_cell.value, str) and summary_cell.value.strip():
            normalized=normalize_ai_summary(summary_cell.value)
            if normalized and normalized != summary_cell.value:
                summary_cell.value=normalized
                cleaned+=1
    return cleaned

def migrate_and_repair_workbook(wb):
    migrated=0
    repaired=0
    ai_cleaned=0
    for sheet_name in wb.sheetnames:
        ws=wb[sheet_name]
        # No tocar hojas vacías temporales
        if ws.max_row==1 and ws.max_column==1 and ws["A1"].value is None:
            continue
        if migrate_sheet_headers(ws):
            migrated+=1
        repaired+=repair_invalid_hyperlinks_in_sheet(ws)
        ai_cleaned += repair_ai_columns_in_sheet(ws)
        if ws.max_row>1:
            apply_table_style(ws, ws.max_row)
    if migrated or repaired or ai_cleaned:
        print(
            "[INFO] Migración/Reparación aplicada - "
            f"hojas migradas: {migrated}, hipervínculos reparados: {repaired}, "
            f"resúmenes IA limpiados: {ai_cleaned}"
        )
    return migrated, repaired, ai_cleaned

def append_rows(wb,df):
    if df.empty:
        return
    
    hoja = get_month_name_from_period(df)
    
    # Verificar si la hoja ya existe
    if hoja not in wb.sheetnames:
        ws=wb.create_sheet(title=hoja)
        ws.append(expected_columns)
        # Aplicar estilo inmediatamente al crear nueva hoja
        apply_table_style(ws, 1)
        print(f"[INFO] Nueva hoja '{hoja}' creada con estilo")
    else:
        ws=wb[hoja]
        if migrate_sheet_headers(ws):
            apply_table_style(ws, 1)
            print(f"[INFO] Hoja '{hoja}' migrada al nuevo esquema de columnas")
    
    # Obtener claves existentes para verificar duplicados (preferimos URL del mensaje de Slack)
    existing_keys = set()
    slack_col=get_column_index(ws, "SLACK", fallback=2)
    if ws.max_row > 1:  # Si hay datos además del header
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if slack_col is not None and len(row) >= slack_col and row[slack_col-1]:
                slack_content = str(row[slack_col-1]).strip()
                url = extract_hyperlink_url(slack_content)
                existing_keys.add(url or slack_content)
    
    # Agregar solo mensajes nuevos
    new_rows_added = 0
    for _,r in df.iterrows():
        slack_content = str(r["SLACK"]).strip()
        if slack_content:
            key = extract_hyperlink_url(slack_content) or slack_content
            if key and key not in existing_keys:
                ws.append(build_row_values_for_sheet(ws, r.to_dict()))
                existing_keys.add(key)
                new_rows_added += 1
    
    print(f"[INFO] Filas nuevas agregadas: {new_rows_added} (duplicados ignorados: {len(df) - new_rows_added})")
    
    # Aplicar estilo a la tabla siempre (incluso si no hay filas nuevas)
    if ws.max_row > 1:  # Si hay datos además del header
        apply_table_style(ws, ws.max_row)
        print(f"[INFO] Estilo aplicado a la tabla")
    
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row==1 and wb["Sheet"].max_column==1 and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])

def main():
    print(f"[INFO] Inicio ejecución: {now_scl()}")
    get_blackbox_guide_text()
    print(f"[INFO] Clasificación por guía .md estricta: {strict_md_classification}")

    token=None
    if local_mode:
        print(f"[INFO] Modo ejecución: {run_mode} | Excel local: {os.path.abspath(local_excel_path)}")
        ensure_local_file(local_excel_path)
        bio=dl_excel_local(local_excel_path)
    else:
        if not onedrive_upn:
            raise RuntimeError("Falta ONEDRIVE_UPN para ejecución con OneDrive")
        print(f"[INFO] Modo ejecución: {run_mode} | OneDrive destino: {onedrive_file_path}")
        token=acquire_token()
        print("[INFO] Access token obtenido")
        ensure_file(token)
        bio=dl_excel(token)

    try:
        wb=load_workbook(bio)
        print("[INFO] Excel cargado")
    except Exception:
        wb=Workbook()
        wb.active.title="TMP"
        print("[WARN] Excel nuevo creado")

    migrated_sheets, repaired_links, ai_cleaned = migrate_and_repair_workbook(wb)
    existing_slack_keys=collect_existing_slack_keys(wb)
    existing_row_locations=collect_existing_row_locations(wb)
    print(f"[INFO] Claves Slack ya existentes en Excel: {len(existing_slack_keys)}")

    # Ejecutar siempre, sin restricción de hora
    print(f"[INFO] Ejecutando sin restricción de hora (debug_mode: {debug_mode})")

    # Ventana Slack configurable por días hacia atrás (en horario Chile).
    # Para pruebas locales, por defecto usa solo hoy (0 días).
    lookback_days_default="0" if local_mode else "3"
    lookback_days_raw=os.environ.get("SLACK_LOOKBACK_DAYS", lookback_days_default).strip() or lookback_days_default
    try:
        lookback_days=max(0, int(lookback_days_raw))
    except ValueError:
        print(f"[WARN] SLACK_LOOKBACK_DAYS inválido: {lookback_days_raw}. Usando {lookback_days_default}.")
        lookback_days=int(lookback_days_default)

    now_local = now_scl()
    start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=lookback_days)
    oldest = str(start_local.astimezone(timezone.utc).timestamp())
    latest = str(datetime.now(tz=timezone.utc).timestamp())
    if lookback_days == 0:
        print(f"[INFO] Ventana Slack: solo hoy (hora Chile): {start_local} hasta {now_local}")
    else:
        print(f"[INFO] Ventana Slack: últimos {lookback_days + 1} días (hora Chile): {start_local} hasta {now_local}")
    msgs = fetch_messages(oldest=oldest, latest=latest)

    print(f"[INFO] Mensajes obtenidos: {len(msgs)}")
    enable_backfill=(os.environ.get("ENABLE_BACKFILL","1").strip()=="1")
    if enable_backfill:
        _, backfilled_ai, backfilled_status, backfilled_classification = backfill_ai_for_existing_rows(wb, msgs, existing_row_locations)
    else:
        backfilled_ai, backfilled_status, backfilled_classification = 0, 0, 0
        print("[INFO] Backfill sobre filas existentes desactivado (ENABLE_BACKFILL=0)")
    df=build_df(msgs, existing_keys=existing_slack_keys)
    workbook_changed = (
        migrated_sheets > 0
        or repaired_links > 0
        or ai_cleaned > 0
        or backfilled_ai > 0
        or backfilled_status > 0
        or backfilled_classification > 0
    )
    if not df.empty:
        print(f"[INFO] Filas a agregar: {len(df)}")
        print("[DEBUG] Preview:\n", df.head(5).to_string())
        append_rows(wb,df)
        sheet_name = get_month_name_from_period(df)
        print(f"[INFO] Datos procesados en hoja '{sheet_name}'")
        workbook_changed = True
    else:
        print("[INFO] No hay mensajes nuevos")
        if not workbook_changed:
            return

    out=io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    if local_mode:
        upload_success = up_excel_local(local_excel_path, out)
        if upload_success:
            print(f"[INFO] Excel actualizado en local: {os.path.abspath(local_excel_path)}")
        else:
            print("[WARN] No se pudo guardar Excel local, pero el procesamiento se completó")
    else:
        upload_success = up_excel(token,out)
        if upload_success:
            print(f"[INFO] Excel actualizado en OneDrive: {onedrive_file_path}")
        else:
            print(f"[WARN] No se pudo actualizar OneDrive, pero el procesamiento se completó exitosamente")

    print(f"[INFO] Fin ejecución: {now_scl()}")

if __name__=="__main__":
    main()
